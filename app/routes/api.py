from flask import Blueprint, jsonify, request, session, Response
from datetime import datetime, timedelta
from ..db import get_spar_db
import pytz
import csv
import io
import math
import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

api_bp = Blueprint("api", __name__)

IST = pytz.timezone("Asia/Kolkata")

def login_required_api(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated

def get_date_range():
    """Parse from/to query params (YYYY-MM-DD), default last 30 days."""
    now = datetime.now(IST)
    default_from = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    default_to = now.strftime("%Y-%m-%d")
    date_from = request.args.get("from", default_from)
    date_to   = request.args.get("to",   default_to)
    # Add one day to to_date to make it inclusive
    try:
        dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
        date_to_exclusive = dt_to.strftime("%Y-%m-%d")
    except Exception:
        date_to_exclusive = date_to
    return date_from, date_to, date_to_exclusive

def str_date_filter(from_str, to_exclusive_str, field="date_time"):
    """Return a MongoDB filter for string-based date_time fields."""
    return {field: {"$gte": from_str, "$lt": to_exclusive_str}}


def get_hour_range():
    """Parse hour_from/hour_to query params, but strictly bound to 9 AM – 10 PM."""
    hf = int(request.args.get("hour_from", 9))
    ht = int(request.args.get("hour_to", 22))
    hour_from = max(9, min(22, hf))
    hour_to   = max(9, min(22, ht))
    if hour_from > hour_to:
        hour_from, hour_to = 9, 22
    return hour_from, hour_to


def hour_expr_str(hour_from, hour_to):
    """$expr hour filter for string date_time fields."""
    if hour_from == 0 and hour_to == 23:
        return {}
    return {"$expr": {"$and": [
        {"$gte": [{"$substr": ["$date_time", 11, 2]}, f"{hour_from:02d}"]},
        {"$lte": [{"$substr": ["$date_time", 11, 2]}, f"{hour_to:02d}"]},
    ]}}

def get_store_code():
    """Parse store_code query param, default to first store."""
    store_code = request.args.get("store_code", "").strip()
    if store_code:
        return store_code
    return "Spar-20016-TSM-Mall-Udupi"  # Default store

def get_pin_filters():
    """Parse cross-chart pin filter params."""
    pin_date   = request.args.get("pin_date",   "").strip()
    pin_hour   = request.args.get("pin_hour",   "").strip()
    pin_gender = request.args.get("pin_gender", "").strip()
    pin_age    = request.args.get("pin_age",    "").strip()
    try:
        ph = int(pin_hour) if pin_hour != "" else None
    except (ValueError, TypeError):
        ph = None
    return (pin_date or None), ph, (pin_gender or None), (pin_age or None)

def gender_count_expr(pin_gender=None):
    """MongoDB $sum expression for visitor counts, optionally filtered to one gender."""
    def conv(f):
        return {"$convert": {"input": f"${f}", "to": "int", "onError": 0, "onNull": 0}}
    if pin_gender == "male":   return conv("count_male")
    if pin_gender == "female": return conv("count_female")
    if pin_gender == "child":  return conv("count_child")
    return {"$add": [conv("count_male"), conv("count_female"), conv("count_child")]}

def pin_date_filt(pin_date):
    """Build a one-day str_date_filter from a pin_date string."""
    next_d = (datetime.strptime(pin_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    return str_date_filter(pin_date, next_d)

def get_today_range():
    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")
    return today, tomorrow

def get_yesterday_range():
    now = datetime.now(IST)
    yesterday = (now - timedelta(days=1)).strftime("%Y-%m-%d")
    today = now.strftime("%Y-%m-%d")
    return yesterday, today

def get_this_week_range():
    now = datetime.now(IST)
    week_start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")
    return week_start, tomorrow

def get_last_week_range():
    now = datetime.now(IST)
    week_start = now - timedelta(days=now.weekday())
    last_week_start = (week_start - timedelta(weeks=1)).strftime("%Y-%m-%d")
    last_week_end = week_start.strftime("%Y-%m-%d")
    return last_week_start, last_week_end

def get_this_month_range():
    now = datetime.now(IST)
    month_start = now.replace(day=1).strftime("%Y-%m-%d")
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")
    return month_start, tomorrow

def get_last_month_range():
    now = datetime.now(IST)
    this_month_start = now.replace(day=1)
    last_month_end = this_month_start.strftime("%Y-%m-%d")
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1).strftime("%Y-%m-%d")
    return last_month_start, last_month_end

def footfall_sum(col, from_str, to_exclusive_str, store_code=None):
    """Aggregate visitor footfall totals from sparServer.footfall."""
    def conv(f): return {"$convert": {"input": f"${f}", "to": "int", "onError": 0, "onNull": 0}}
    match_filter = str_date_filter(from_str, to_exclusive_str)
    if store_code:
        match_filter["store_code"] = store_code
    pipeline = [
        {"$match": match_filter},
        {"$group": {"_id": None, "total": {"$sum": {"$add": [
            conv("count_male"), conv("count_female"), conv("count_child")
        ]}}}},
    ]
    res = list(col.aggregate(pipeline))
    return res[0]["total"] if res else 0

# ─────────────────────────────────────────────
#  OVERVIEW
# ─────────────────────────────────────────────
@api_bp.route("/overview")
@login_required_api
def overview():
    ff_col = get_spar_db().footfall

    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    # Period comparison — visitor counts (no hour filter on chips)
    periods = {}
    for label, (f, t) in [
        ("today",      get_today_range()),
        ("yesterday",  get_yesterday_range()),
        ("this_week",  get_this_week_range()),
        ("last_week",  get_last_week_range()),
        ("this_month", get_this_month_range()),
        ("last_month", get_last_month_range()),
    ]:
        periods[label] = footfall_sum(ff_col, f, t, store_code)

    pin_date, pin_hour, pin_gender, pin_age = get_pin_filters()

    # Effective date+hour filter (pin overrides range)
    bd_filt = pin_date_filt(pin_date) if pin_date else str_date_filter(date_from, date_to_ex)
    bd_filt["store_code"] = store_code
    eff_h_from = pin_hour if pin_hour is not None else hour_from
    eff_h_to   = pin_hour if pin_hour is not None else hour_to
    bd_filt.update(hour_expr_str(eff_h_from, eff_h_to))

    if pin_age:
        # Gender breakdown from age_group collection for that age group
        age_docs = list(get_spar_db().age_group.aggregate([
            {"$match": {"age_group": pin_age, **bd_filt}},
            {"$group": {"_id": "$gender", "count": {"$sum": 1}}},
        ]))
        breakdown = {"male": 0, "female": 0, "child": 0, "staff": 0}
        for d in age_docs:
            if d["_id"] in breakdown:
                breakdown[d["_id"]] = d["count"]
    else:
        def conv(f): return {"$convert": {"input": f"${f}", "to": "int", "onError": 0, "onNull": 0}}
        ff = list(ff_col.aggregate([
            {"$match": bd_filt},
            {"$group": {
                "_id":    None,
                "male":   {"$sum": conv("count_male")},
                "female": {"$sum": conv("count_female")},
                "child":  {"$sum": conv("count_child")},
                "staff":  {"$sum": conv("count_staff")},
            }}
        ]))
        breakdown = ff[0] if ff else {"male": 0, "female": 0, "child": 0, "staff": 0}
        breakdown.pop("_id", None)
    total_visitors = breakdown.get("male", 0) + breakdown.get("female", 0) + breakdown.get("child", 0)

    return jsonify({
        "periods":        periods,
        "breakdown":      breakdown,
        "total_visitors": total_visitors,
        "date_from":      date_from,
        "date_to":        date_to,
    })

# ─────────────────────────────────────────────
#  DAILY FOOTFALL TREND
# ─────────────────────────────────────────────
@api_bp.route("/trend")
@login_required_api
def trend():
    col = get_spar_db().footfall
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    pin_date, pin_hour, pin_gender, _ = get_pin_filters()
    vis_expr = gender_count_expr(pin_gender)
    stf_expr = {"$convert": {"input": "$count_staff", "to": "int", "onError": 0, "onNull": 0}}

    if pin_date:
        # Switch to hourly mode: show 24 hourly slots for that specific date
        filt = pin_date_filt(pin_date)
        filt["store_code"] = store_code
        if pin_hour is not None:
            filt.update(hour_expr_str(pin_hour, pin_hour))
        group_id = {"$substr": ["$date_time", 11, 2]}   # "HH"
        mode = "hourly"
    else:
        filt = str_date_filter(date_from, date_to_ex)
        filt["store_code"] = store_code
        h_from = pin_hour if pin_hour is not None else hour_from
        h_to   = pin_hour if pin_hour is not None else hour_to
        filt.update(hour_expr_str(h_from, h_to))
        group_id = {"$substr": ["$date_time", 0, 10]}   # "YYYY-MM-DD"
        mode = "daily"

    rows = list(col.aggregate([
        {"$match": filt},
        {"$group": {"_id": group_id, "visitors": {"$sum": vis_expr}, "staff": {"$sum": stf_expr}}},
        {"$sort": {"_id": 1}}
    ]))

    labels   = [r["_id"]      for r in rows]
    visitors = [r["visitors"] for r in rows]
    staff    = [r["staff"]    for r in rows]
    return jsonify({"labels": labels, "visitors": visitors, "staff": staff, "mode": mode})

# ─────────────────────────────────────────────
#  HOURLY FOOTFALL
# ─────────────────────────────────────────────
@api_bp.route("/hourly")
@login_required_api
def hourly():
    col = get_spar_db().footfall
    date_from, date_to, date_to_ex = get_date_range()
    store_code = get_store_code()
    pin_date, _ph, pin_gender, _ = get_pin_filters()
    vis_expr = gender_count_expr(pin_gender)
    hour_from, hour_to = get_hour_range()
    hour_map = {f"{h:02d}": 0 for h in range(hour_from, hour_to + 1)}

    if pin_date:
        # Actual counts for that specific date (not average)
        filt = {**pin_date_filt(pin_date), **hour_expr_str(hour_from, hour_to), "store_code": store_code}
        rows = list(col.aggregate([
            {"$match": filt},
            {"$group": {"_id": {"$substr": ["$date_time", 11, 2]}, "total": {"$sum": vis_expr}}},
            {"$sort": {"_id": 1}}
        ]))
        for r in rows:
            if r["_id"] in hour_map:
                hour_map[r["_id"]] = r["total"]
        is_avg = False
    else:
        # Average across date range
        rows = list(col.aggregate([
            {"$match": {**str_date_filter(date_from, date_to_ex), **hour_expr_str(hour_from, hour_to), "store_code": store_code}},
            {"$group": {
                "_id":   {"$substr": ["$date_time", 11, 2]},
                "total": {"$sum": vis_expr},
                "dates": {"$addToSet": {"$substr": ["$date_time", 0, 10]}},
            }},
            {"$sort": {"_id": 1}}
        ]))
        for r in rows:
            if r["_id"] in hour_map:
                n = len(r["dates"])
                hour_map[r["_id"]] = round(r["total"] / n) if n else 0
        is_avg = True

    labels = [f"{h:02d}:00" for h in range(hour_from, hour_to + 1)]
    values = [hour_map[f"{h:02d}"] for h in range(hour_from, hour_to + 1)]
    return jsonify({"labels": labels, "values": values,
                    "date_from": date_from, "date_to": date_to,
                    "pin_date": pin_date, "is_avg": is_avg})

# ─────────────────────────────────────────────
#  DEVICE STATUS (derived from last data received)
# ─────────────────────────────────────────────
@api_bp.route("/devices")
@login_required_api
def devices():
    db = get_spar_db()
    now = datetime.now(IST)
    store_code = get_store_code()

    results = []
    for col_name, label in [("footfall", "Footfall Camera"), ("heatmap", "Heatmap Camera"), ("queue_length", "Queue Length Camera")]:
        col = db[col_name]
        total = col.count_documents({"store_code": store_code})
        if total == 0:
            continue

        # queue_length collection may not have device_serial_id; use camera_no instead
        if col_name == "queue_length":
            devices_in_col = col.distinct("camera_no", {"store_code": store_code})
            id_field = "camera_no"
        else:
            devices_in_col = col.distinct("device_serial_id", {"store_code": store_code})
            id_field = "device_serial_id"

        for dev_id in devices_in_col:
            last_doc = col.find_one({id_field: dev_id, "store_code": store_code}, sort=[("date_time", -1)])
            if not last_doc:
                continue
            last_seen_str = last_doc.get("date_time","")
            camera = last_doc.get("camera_no","")
            store  = last_doc.get("store_code","")

            # For queue_length, use camera_no as display device_id if no device_serial_id
            display_id = last_doc.get("device_serial_id", "") if col_name != "queue_length" else (last_doc.get("device_serial_id", "") or f"Camera {dev_id}")

            # Parse last seen — handle both "YYYY-MM-DD HH:MM:SS" and "YYYY-MM-DD_HH-MM-SS" formats
            try:
                clean_str = last_seen_str[:19].replace("_", " ").replace("-", ":", 2)
                # Fix: only replace hyphens in time portion, not date
                date_part = last_seen_str[:10]
                time_part = last_seen_str[11:19].replace("-", ":") if len(last_seen_str) >= 19 else ""
                parse_str = f"{date_part} {time_part}" if time_part else date_part
                last_seen_dt = datetime.strptime(parse_str, "%Y-%m-%d %H:%M:%S")
                last_seen_dt = IST.localize(last_seen_dt)
                diff_minutes = (now - last_seen_dt).total_seconds() / 60
                if diff_minutes < 20:
                    status = "online"
                elif diff_minutes < 120:
                    status = "warning"
                else:
                    status = "offline"
                last_seen_friendly = last_seen_str[:16]
            except Exception:
                status = "offline"
                last_seen_friendly = last_seen_str[:16]
                diff_minutes = 9999

            results.append({
                "device_id": display_id,
                "collection": col_name,
                "label": label,
                "camera": camera,
                "store": store,
                "last_seen": last_seen_friendly,
                "status": status,
                "minutes_ago": round(diff_minutes),
            })

    return jsonify({"devices": results})

# ─────────────────────────────────────────────
#  EXPORT FOOTFALL EXCEL
# ─────────────────────────────────────────────
@api_bp.route("/export-footfall")
@login_required_api
def export_footfall():
    col = get_spar_db().footfall
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    docs = list(col.find(
        {**str_date_filter(date_from, date_to_ex), **hour_expr_str(hour_from, hour_to), "store_code": store_code}
    ).sort("date_time", -1).limit(5000))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Footfall"

    header_fill = PatternFill("solid", fgColor="DA291C")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["Date/Time", "Male", "Female", "Child", "Staff", "Total Visitors"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, d in enumerate(docs, 2):
        m  = int(d.get("count_male",  0) or 0)
        f_ = int(d.get("count_female",0) or 0)
        c  = int(d.get("count_child", 0) or 0)
        s  = int(d.get("count_staff", 0) or 0)
        ws.cell(row=ri, column=1, value=d.get("date_time", ""))
        ws.cell(row=ri, column=2, value=m)
        ws.cell(row=ri, column=3, value=f_)
        ws.cell(row=ri, column=4, value=c)
        ws.cell(row=ri, column=5, value=s)
        ws.cell(row=ri, column=6, value=m + f_ + c)

    for i, col_dim in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename=spar_footfall_{date_from}_{date_to}.xlsx"}
    )

# ─────────────────────────────────────────────
#  HEATMAP DATA
# ─────────────────────────────────────────────

# Dynamic location discovery based on available images
import os

def get_location_map(store_code):
    """Dynamically build location map from available images in store folder."""
    from pathlib import Path
    import re

    store_folder = Path(__file__).parent.parent / "static" / "img" / "heatmap" / store_code
    location_map = {}

    if not store_folder.exists():
        return {}

    # Find all .jpg files and extract location info
    for img_file in sorted(store_folder.glob("location*.jpg")):
        filename = img_file.stem  # filename without .jpg
        # Extract location code (e.g., "location01" from "location01- FMCG Food")
        match = re.match(r"(location\d+)", filename)
        if match:
            location_id = match.group(1)
            # Extract label (everything after " - ")
            if " - " in filename:
                label = filename.split(" - ", 1)[1]
            else:
                label = filename.replace(location_id, "").strip(" - ")

            location_map[location_id] = {
                "image": filename,
                "label": label or location_id
            }

    return location_map

# Legacy static map for backward compatibility
LOCATION_MAP = {
    "location01": {"image": "location01- FMCG Food",      "label": "FMCG Food"},
    "location02": {"image": "location02- Grocery",         "label": "Grocery"},
    "location03": {"image": "location03- Grocery 2",       "label": "Grocery 2"},
    "location04": {"image": "location04- Home and Living", "label": "Home and Living"},
    "location05": {"image": "location05- FMCG Food 2",     "label": "FMCG Food 2"},
    "location06": {"image": "location06- GM- Stationery",  "label": "GM- Stationery"},
    "location07": {"image": "location07- Dairy and Frozen","label": "Dairy and Frozen"},
    "location08": {"image": "location08- Cash Counter",    "label": "Cash Counter"},
    "location09": {"image": "location09- FMCG Non Food",   "label": "FMCG Non Food"},
}

@api_bp.route("/heatmap-data")
@login_required_api
def heatmap_data():
    from collections import OrderedDict
    location  = request.args.get("location", "location01")
    hour_from = int(request.args.get("hour_from", 0))
    hour_to   = int(request.args.get("hour_to",  23))
    store_code = get_store_code()

    # Get dynamic location map for this store
    location_map = get_location_map(store_code) or LOCATION_MAP

    if location not in location_map:
        return jsonify({"error": "Unknown location"}), 400

    col = get_spar_db().heatmap

    # Accept comma-separated dates or single date (backwards compat)
    dates_param = request.args.get("dates", "") or request.args.get("date", "")
    dates = [d.strip() for d in dates_param.split(",") if d.strip()]

    if not dates:
        latest = col.find_one({"store_location": location, "store_code": store_code}, sort=[("date_time", -1)])
        dates = [latest["date_time"][:10]] if latest else []
        if not dates:
            return jsonify({"frames": [], "total": {"male": 0, "female": 0, "child": 0, "staff": 0}})

    # Build per-date hour-range conditions
    or_conditions = []
    for d in dates:
        tf = f"{d} {hour_from:02d}:00:00"
        tt_h = hour_to + 1
        if tt_h >= 24:
            next_d = (datetime.strptime(d, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
            tt = f"{next_d} 00:00:00"
        else:
            tt = f"{d} {tt_h:02d}:00:00"
        or_conditions.append({"date_time": {"$gte": tf, "$lt": tt}})

    if len(or_conditions) == 1:
        filt = {"store_location": location, "store_code": store_code, **or_conditions[0]}
    else:
        filt = {"store_location": location, "store_code": store_code, "$or": or_conditions}

    docs = list(col.find(
        filt,
        {"date_time": 1, "count": 1, "person_bbox_list": 1, "_id": 0}
    ).sort("date_time", 1))

    # Group by HH:MM — merge bboxes & counts across all selected dates
    slots  = OrderedDict()   # keyed by HH:MM
    total  = {"male": 0, "female": 0, "child": 0, "staff": 0}
    multi  = len(dates) > 1

    for doc in docs:
        ts = doc["date_time"][11:16]   # HH:MM
        if ts not in slots:
            slots[ts] = {
                "datetime": ts if multi else doc["date_time"],
                "time":     ts,
                "count":    {"male": 0, "female": 0, "child": 0, "staff": 0},
                "bboxes":   {"male": [], "female": [], "child": [], "staff": []},
            }
        slot = slots[ts]

        cnt = doc.get("count") or {}
        if isinstance(cnt, str):
            import ast
            try: cnt = ast.literal_eval(cnt)
            except: cnt = {}

        bbl = doc.get("person_bbox_list") or {}
        if isinstance(bbl, str):
            import ast
            try: bbl = ast.literal_eval(bbl)
            except: bbl = {}

        for cat in ("male", "female", "child", "staff"):
            raw  = bbl.get(cat, []) or []
            slot["bboxes"][cat].extend([b for b in raw if isinstance(b, list) and len(b) == 4])
            c = int(cnt.get(cat, 0))
            slot["count"][cat] += c
            total[cat] += c

        # Keep full datetime label for single-date view
        if not multi:
            slot["datetime"] = doc["date_time"]

    loc_info = location_map[location]
    # Build store-specific image path
    store_image = f"/static/img/heatmap/{store_code}/{loc_info['image']}.jpg"
    return jsonify({
        "location":  location,
        "image":     store_image,
        "label":     loc_info["label"],
        "dates":     dates,
        "hour_from": hour_from,
        "hour_to":   hour_to,
        "frames":    list(slots.values()),
        "total":     total,
        "locations": [{"id": k, "label": v["label"]} for k, v in location_map.items()],
    })

@api_bp.route("/heatmap-table")
@login_required_api
def heatmap_table():
    """Return per-section, per-hour occupied minutes for the tabular heatmap view.

    Each cell = number of distinct minute-slots (HH:MM) in that hour where at least
    one person was detected, averaged across the selected dates.  Max value = 60.
    """
    from collections import defaultdict
    import ast

    dates_param = request.args.get("dates", "")
    dates = [d.strip() for d in dates_param.split(",") if d.strip()]
    hour_from = int(request.args.get("hour_from", 9))
    hour_to   = int(request.args.get("hour_to",  22))
    store_code = get_store_code()

    if not dates:
        return jsonify({"error": "No dates provided"}), 400

    # Use dynamic location map for this store (falls back to legacy map)
    location_map = get_location_map(store_code) or LOCATION_MAP

    col = get_spar_db().heatmap

    or_conditions = []
    for d in dates:
        tf   = f"{d} {hour_from:02d}:00:00"
        tt_h = hour_to + 1
        if tt_h >= 24:
            next_d = (datetime.strptime(d, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
            tt = f"{next_d} 00:00:00"
        else:
            tt = f"{d} {tt_h:02d}:00:00"
        or_conditions.append({"date_time": {"$gte": tf, "$lt": tt}})

    filt = or_conditions[0] if len(or_conditions) == 1 else {"$or": or_conditions}
    filt["store_code"] = store_code

    docs = list(col.find(filt, {"date_time": 1, "store_location": 1, "count": 1, "_id": 0}))

    hours     = [f"{h:02d}" for h in range(hour_from, hour_to + 1)]
    hours_set = set(hours)

    # presence[loc][hour_str][date_str] = set of "HH:MM" minute-strings with detections
    presence = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))
    # counts[loc][hour_str] = cumulative detection counts across all docs
    counts = defaultdict(lambda: defaultdict(lambda: {"male": 0, "female": 0, "child": 0, "staff": 0}))

    for doc in docs:
        dt_str = doc.get("date_time", "")
        if len(dt_str) < 16:
            continue
        date_str   = dt_str[:10]       # YYYY-MM-DD
        hour_str   = dt_str[11:13]     # HH
        minute_str = dt_str[11:16]     # HH:MM

        if hour_str not in hours_set:
            continue
        loc = doc.get("store_location", "")
        if loc not in location_map:
            continue

        cnt = doc.get("count") or {}
        if isinstance(cnt, str):
            try: cnt = ast.literal_eval(cnt)
            except: cnt = {}

        cats = {cat: int(cnt.get(cat, 0) or 0) for cat in ("male", "female", "child", "staff")}
        total = sum(cats.values())
        if total > 0:
            presence[loc][hour_str][date_str].add(minute_str)
        for cat, v in cats.items():
            counts[loc][hour_str][cat] += v

    # Build table: average occupied-minutes across selected dates, cap at 60
    n_dates = len(dates)
    table = {}
    for h in hours:
        table[h] = {}
        for loc in location_map:
            hour_data = presence[loc].get(h, {})
            total_mins = sum(len(mins) for mins in hour_data.values())
            avg_mins   = round(total_mins / n_dates) if n_dates else 0
            c = counts[loc].get(h, {"male": 0, "female": 0, "child": 0, "staff": 0})
            table[h][loc] = {
                "minutes": min(60, avg_mins),
                "male":    c["male"],
                "female":  c["female"],
                "child":   c["child"],
                "staff":   c["staff"],
            }

    sections    = [{"id": k, "label": v["label"]} for k, v in location_map.items()]
    hour_labels = [f"{h}:00" for h in hours]

    return jsonify({
        "hours":       hours,
        "hour_labels": hour_labels,
        "sections":    sections,
        "table":       table,
        "n_dates":     n_dates,
    })


@api_bp.route("/heatmap-locations")
@login_required_api
def heatmap_locations():
    """Return available locations/cameras for the selected store."""
    store_code = get_store_code()

    # Get dynamic location map for this store
    location_map = get_location_map(store_code) or LOCATION_MAP

    cameras = [
        {"camera_id": loc_id, "name": loc_info["label"]}
        for loc_id, loc_info in location_map.items()
    ]
    return jsonify({"cameras": cameras})

@api_bp.route("/heatmap-dates")
@login_required_api
def heatmap_dates():
    """Return distinct dates available per store location."""
    location = request.args.get("location", "location01")
    store_code = get_store_code()

    # Get dynamic location map for this store
    location_map = get_location_map(store_code) or LOCATION_MAP

    if location not in location_map:
        return jsonify({"dates": []})
    col  = get_spar_db().heatmap
    docs = list(col.find({"store_location": location, "store_code": store_code}, {"date_time": 1, "_id": 0}))
    dates = sorted({d["date_time"][:10] for d in docs})
    return jsonify({"dates": dates})

# ─────────────────────────────────────────────
#  QUEUE & WAIT ANALYSIS (sparServer.queue_length)
# ─────────────────────────────────────────────
@api_bp.route("/queue-stats")
@login_required_api
def queue_stats():
    db = get_spar_db()
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()
    col = db.queue_length

    # date_time format: "YYYY-MM-DD_HH-MM-SS"
    # str_date_filter still works — lexicographic ordering is preserved
    filt = str_date_filter(date_from, date_to_ex)
    filt["store_code"] = store_code
    filt.update(hour_expr_str(hour_from, hour_to))

    docs = list(col.find(filt, {
        "camera_no": 1, "date_time": 1,
        "avg_wait_time": 1, "person_detections": 1,
        "_id": 0
    }))

    # Per-camera accumulators
    cam_stats = {}
    for doc in docs:
        cam = doc.get("camera_no")
        if cam is None:
            continue
        if cam not in cam_stats:
            cam_stats[cam] = {
                "wait_times":   [],
                "queue_lengths":[],
                "hourly_wait":  {},   # int hour -> [values]
                "hourly_queue": {},
            }
        cs = cam_stats[cam]

        # Wait time already in minutes — average non-zero
        wait = doc.get("avg_wait_time")
        if wait and isinstance(wait, (int, float)) and wait > 0:
            cs["wait_times"].append(float(wait))

        # Queue length: flatten person_detections list-of-lists, avg non-zero
        pdet = doc.get("person_detections") or []
        flat = []
        for sublist in pdet:
            if isinstance(sublist, list):
                flat.extend(v for v in sublist if isinstance(v, (int, float)) and v > 0)
        ql = (sum(flat) / len(flat)) if flat else None
        if ql is not None:
            # Stochastic rounding: randomly ceil or floor weighted by fractional part
            frac = ql - math.floor(ql)
            ql_rounded = math.ceil(ql) if random.random() < frac else math.floor(ql)
            ql_rounded = max(ql_rounded, 1)  # ensure at least 1 if there was a detection
            cs["queue_lengths"].append(float(ql_rounded))

        # Hourly bucket: date_time[11:13] = "HH" for "YYYY-MM-DD_HH-MM-SS"
        dt_str = doc.get("date_time", "")
        if len(dt_str) >= 13:
            try:
                h = int(dt_str[11:13])
                if wait and wait > 0:
                    cs["hourly_wait"].setdefault(h, []).append(float(wait))
                if ql is not None:
                    cs["hourly_queue"].setdefault(h, []).append(ql)
            except ValueError:
                pass

    # Build per-camera summary list
    cameras = []
    for cam_no in sorted(cam_stats.keys()):
        cs  = cam_stats[cam_no]
        wt  = cs["wait_times"]
        qlv = cs["queue_lengths"]
        cameras.append({
            "camera_no":    cam_no,
            "label":        f"Counter {cam_no}",
            "avg_wait_min": round(sum(wt)  / len(wt),  2) if wt  else 0,
            "avg_queue":    round(sum(qlv) / len(qlv)) if qlv else 0,
            "max_queue":    round(max(qlv))            if qlv else 0,
            "observations": len(wt),
        })

    # Overall KPIs
    all_waits  = [w for cs in cam_stats.values() for w in cs["wait_times"]]
    all_queues = [q for cs in cam_stats.values() for q in cs["queue_lengths"]]
    overall_avg_wait  = round(sum(all_waits)  / len(all_waits),  2) if all_waits  else 0
    overall_avg_queue = round(sum(all_queues) / len(all_queues)) if all_queues else 0
    active_counters   = len([c for c in cameras if c["observations"] > 0])

    # Global hourly averages (all counters combined)
    g_hw, g_hq = {}, {}
    for cs in cam_stats.values():
        for h, vals in cs["hourly_wait"].items():
            g_hw.setdefault(h, []).extend(vals)
        for h, vals in cs["hourly_queue"].items():
            g_hq.setdefault(h, []).extend(vals)

    hourly_labels = [f"{h:02d}:00" for h in range(hour_from, hour_to + 1)]
    hourly_wait_vals  = [
        round(sum(g_hw[h]) / len(g_hw[h]), 2) if g_hw.get(h) else None
        for h in range(hour_from, hour_to + 1)
    ]
    hourly_queue_vals = [
        round(sum(g_hq[h]) / len(g_hq[h])) if g_hq.get(h) else None
        for h in range(hour_from, hour_to + 1)
    ]

    # Per-counter hourly breakdown (for multi-line charts)
    per_counter_hourly = []
    for cam_no in sorted(cam_stats.keys()):
        cs = cam_stats[cam_no]
        hw, hq = [], []
        for h in range(hour_from, hour_to + 1):
            wvals = cs["hourly_wait"].get(h, [])
            qvals = cs["hourly_queue"].get(h, [])
            hw.append(round(sum(wvals) / len(wvals), 2) if wvals else None)
            hq.append(round(sum(qvals) / len(qvals)) if qvals else None)
        per_counter_hourly.append({
            "camera_no": cam_no,
            "label":     f"Counter {cam_no}",
            "wait":      hw,
            "queue":     hq,
        })

    return jsonify({
        "cameras":            cameras,
        "overall_avg_wait":   overall_avg_wait,
        "overall_avg_queue":  overall_avg_queue,
        "active_counters":    active_counters,
        "total_observations": len(docs),
        "date_from":          date_from,
        "date_to":            date_to,
        "hourly_labels":      hourly_labels,
        "hourly_wait":        hourly_wait_vals,
        "hourly_queue":       hourly_queue_vals,
        "per_counter_hourly": per_counter_hourly,
    })

# ─────────────────────────────────────────────
#  DWELL TIME (sparServer)
# ─────────────────────────────────────────────
@api_bp.route("/dwell")
@login_required_api
def dwell():
    """Aggregate dwell-time bucket counts from sparServer.dwell_time_summary."""
    db = get_spar_db()
    store_code = get_store_code()
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()

    daily_pipeline = [
        {"$match": {**str_date_filter(date_from, date_to_ex), **hour_expr_str(hour_from, hour_to), "store_code": store_code}},
        {"$group": {
            "_id":   {"$substr": ["$date_time", 0, 10]},
            "lt2":   {"$sum": "$dwell_store_count_less_than_2_minutes"},
            "b2_10": {"$sum": "$dwell_store_count_between_2_to_10_minutes"},
            "gt10":  {"$sum": "$dwell_store_count_more_than_10_minutes"},
        }},
        {"$sort": {"_id": 1}}
    ]
    rows = list(db.dwell_time_summary.aggregate(daily_pipeline))

    labels = [r["_id"]    for r in rows]
    lt2    = [r["lt2"]    for r in rows]
    b2_10  = [r["b2_10"]  for r in rows]
    gt10   = [r["gt10"]   for r in rows]

    return jsonify({
        "labels":       labels,
        "lt2":          lt2,
        "b2_10":        b2_10,
        "gt10":         gt10,
        "total_lt2":    sum(lt2),
        "total_b2_10":  sum(b2_10),
        "total_gt10":   sum(gt10),
    })

# ─────────────────────────────────────────────
#  AGE GROUP (sparServer)
# ─────────────────────────────────────────────
@api_bp.route("/age-group")
@login_required_api
def age_group():
    """Aggregate age-group counts from sparServer.age_group."""
    db = get_spar_db()
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    pin_date, pin_hour, pin_gender, _ = get_pin_filters()

    filt = pin_date_filt(pin_date) if pin_date else str_date_filter(date_from, date_to_ex)
    filt["store_code"] = store_code
    eff_h_from = pin_hour if pin_hour is not None else hour_from
    eff_h_to   = pin_hour if pin_hour is not None else hour_to
    filt.update(hour_expr_str(eff_h_from, eff_h_to))

    VALID_GROUPS = {"Under 18", "18-25", "25-35", "35-45", "45+"}
    filt["age_group"] = {"$in": list(VALID_GROUPS)}
    if pin_gender:
        filt["gender"] = pin_gender

    rows = list(db.age_group.aggregate([
        {"$match": filt},
        {"$group": {"_id": "$age_group", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]))
    LABEL_MAP = {"Under 18": "Under 18", "18-25": "18–25", "25-35": "25–35", "35-45": "35–45", "45+": "45+"}
    buckets = [
        {"key": r["_id"], "label": LABEL_MAP[r["_id"]], "count": r["count"]}
        for r in rows if r["_id"] in VALID_GROUPS
    ]
    return jsonify({"buckets": buckets})
